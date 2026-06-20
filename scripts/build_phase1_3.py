"""
Phase 1-3: C_Penetration シートの実装
  - PARAMシートの L/k/x0 を正しい初期値に更新
  - 名前付き範囲を定義
  - C_Penetration に実績・フィッティング・予測数式を配置
  - 制度イベント（penetration）を段階的に反映する調整済み予測式
  - 成分ごとに折れ線グラフを作成
"""
import csv
from pathlib import Path
from collections import defaultdict

from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "docs" / "sample_data"
XLSX     = BASE_DIR / "output" / "GE_BS_Forecast_Engine.xlsx"

# ── スタイルヘルパー ───────────────────────────────────────────────────
def _font(bold=False, size=11, color="000000", name="Yu Gothic"):
    return Font(bold=bold, size=size, color=color, name=name)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="center", v="center"):
    return Alignment(horizontal=h, vertical=v)

# ── パラメータ定義 ─────────────────────────────────────────────────────
# PARAMシート列E（列番号5）の行番号
# 確認済みレイアウト: 浸透率セクションは行10〜21
PARAM_ROWS = {
    "L":  {"ING01": 10, "ING02": 11, "ING03": 12, "ING04": 13},
    "k":  {"ING01": 14, "ING02": 15, "ING03": 16, "ING04": 17},
    "x0": {"ING01": 18, "ING02": 19, "ING03": 20, "ING04": 21},
}

# ユーザー指定の初期値
INIT_PARAMS = {
    "ING01": {"L": 0.87, "k": 0.10, "x0": -25.0},
    "ING02": {"L": 0.83, "k": 0.10, "x0": -20.0},
    "ING03": {"L": 0.58, "k": 0.13, "x0":  16.0},
    "ING04": {"L": 0.42, "k": 0.11, "x0":  20.0},
}

INGREDIENTS = {
    "ING01": "メトホルミン塩酸塩",
    "ING02": "アトルバスタチン",
    "ING03": "フィルグラスチム",
    "ING04": "インフリキシマブ",
}

# 浸透率系制度イベント（M_Events から抽出済み）
# event_date=2026-06, lag=3, pen_L_delta=0.05, pen_k_delta=0.02
# 予測開始 month_index=41 (2026-06)
EVENT_START_IDX = 41   # 2026-06
EVENT_LAG       = 3
PEN_L_DELTA     = 0.05
PEN_K_DELTA     = 0.02

# 実績期間: 2023-01〜2026-05 (41ヶ月, month_index 0..40)
# 予測期間: 2026-06〜2029-05 (36ヶ月, month_index 41..76)

def _gen_periods(start_ym, n):
    """start_ym='2023-01' からn期間の月文字列リストを生成"""
    y, m = int(start_ym[:4]), int(start_ym[5:7])
    result = []
    for _ in range(n):
        result.append(f"{y:04d}-{m:02d}")
        m += 1
        if m > 12:
            m, y = 1, y + 1
    return result

ACTUAL_PERIODS   = _gen_periods("2023-01", 41)
FORECAST_PERIODS = _gen_periods("2026-06", 36)
ALL_PERIODS      = ACTUAL_PERIODS + FORECAST_PERIODS   # 77件

# ── IQVIA から成分別浸透率を取得 ───────────────────────────────────────
def _load_penetration():
    pen = defaultdict(dict)
    with open(DATA_DIR / "iqvia_market_data.csv", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            pen[r["ingredient_name"]][r["period"]] = float(r["generic_biosimilar_penetration_rate"])
    # 成分名→成分ID のマッピング
    name_to_id = {}
    with open(DATA_DIR / "master_ingredients.csv", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            name_to_id[r["name"]] = r["ingredient_id"]
    result = {}
    for name, periods in pen.items():
        ing_id = name_to_id.get(name)
        if ing_id:
            result[ing_id] = {p: periods[p] for p in ACTUAL_PERIODS if p in periods}
    return result

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ① PARAMシートの L/k/x0 を正しい初期値に更新
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def update_param(wb):
    ws = wb["PARAM"]
    for ing_id, params in INIT_PARAMS.items():
        for param_key, value in params.items():
            row = PARAM_ROWS[param_key][ing_id]
            ws.cell(row=row, column=5, value=value)   # 列E
    print("  PARAM 浸透率パラメータ更新完了")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ② 名前付き範囲を定義（PARAMシート上の浸透率パラメータ）
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def define_named_ranges(wb):
    PARAM_SHEET = "PARAM"
    for param_key in ("L", "k", "x0"):
        for ing_id, row in PARAM_ROWS[param_key].items():
            range_name = f"PARAM_{param_key}_{ing_id}"
            formula = f"'{PARAM_SHEET}'!$E${row}"
            dn = DefinedName(name=range_name, attr_text=formula)
            wb.defined_names[range_name] = dn
    print("  名前付き範囲定義完了")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ③ C_Penetration シートを構築
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 列レイアウト
COL_PERIOD    = 1  # A
COL_IDX       = 2  # B  month_index
COL_ACTUAL    = 3  # C  actual_penetration
COL_LOGISTIC  = 4  # D  logistic_fitted（全期間 base）
COL_FC_EXC    = 5  # E  forecast_excluded（予測期間, base と同値）
COL_FC_ADJ    = 6  # F  forecast_adjusted（予測期間, 制度イベント反映）
COL_NOTES     = 7  # G  備考

# 色定義
C_ING_HDR  = ["1F4E79", "375623", "4A235A", "7B241C"]
C_HDR_ROW  = "1A5276"
C_ACTUAL   = "EBF5FB"
C_FORECAST = "FEF9E7"

N_ACTUAL   = len(ACTUAL_PERIODS)    # 41
N_FORECAST = len(FORECAST_PERIODS)  # 36
N_ALL      = N_ACTUAL + N_FORECAST  # 77
BLOCK_SIZE = 1 + 1 + N_ALL + 2     # ヘッダー行 + 列ヘッダー + データ + 空白2行 = 81

CHART_WIDTH  = 20   # Excelユニット
CHART_HEIGHT = 14


def _logistic_formula(b_cell: str, l_row: int, k_row: int, x0_row: int) -> str:
    """ロジスティック曲線の Excel 数式（base L,k,x0 参照）"""
    L  = f"PARAM!$E${l_row}"
    k  = f"PARAM!$E${k_row}"
    x0 = f"PARAM!$E${x0_row}"
    return f"={L}/(1+EXP(-{k}*({b_cell}-{x0})))"


def _logistic_adjusted_formula(b_cell: str, l_row: int, k_row: int, x0_row: int) -> str:
    """
    制度イベント反映の調整済みロジスティック数式。
    M_Events: event_date=2026-06(idx=41), lag=3, pen_L_delta=0.05, pen_k_delta=0.02
    ramp = MIN(MAX((month_idx - 41)/3, 0), 1)  → 3ヶ月で線形に立ち上がる
    L_adj = L + 0.05 * ramp
    k_adj = k + 0.02 * ramp
    """
    L  = f"PARAM!$E${l_row}"
    k  = f"PARAM!$E${k_row}"
    x0 = f"PARAM!$E${x0_row}"
    ramp = f"MIN(MAX(({b_cell}-{EVENT_START_IDX})/{EVENT_LAG},0),1)"
    L_adj = f"({L}+{PEN_L_DELTA}*{ramp})"
    k_adj = f"({k}+{PEN_K_DELTA}*{ramp})"
    return f"={L_adj}/(1+EXP(-{k_adj}*({b_cell}-{x0})))"


def build_c_penetration(wb):
    ws = wb["C_Penetration"]
    ws.delete_rows(1, ws.max_row)   # 既存の骨格を削除してゼロから構築
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    # 列幅設定
    col_widths = {1: 10, 2: 12, 3: 18, 4: 18, 5: 18, 6: 20, 7: 36}
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # タイトル行
    ws.merge_cells(f"A1:{get_column_letter(COL_NOTES)}1")
    c = ws["A1"]
    c.value     = "C_Penetration  ―  Step2: GE/BS 浸透率予測（ロジスティック曲線）"
    c.font      = _font(bold=True, size=12, color="FFFFFF")
    c.fill      = _fill("4A235A")
    c.alignment = _align()
    ws.row_dimensions[1].height = 24

    # 注釈
    note = ws.cell(row=2, column=1,
        value=("数式: penetration(t) = L / (1 + EXP(-k*(t-x0)))  "
               "| 調整済: LAG=3ヶ月で線形に立ち上がるイベント反映 "
               "| PARAMシートで L/k/x0 を変更すると自動再計算"))
    note.font      = _font(size=9, color="5D5D5D")
    note.alignment = Alignment(horizontal="left")
    ws.merge_cells(f"A2:{get_column_letter(COL_NOTES)}2")
    ws.row_dimensions[2].height = 15

    pen_data = _load_penetration()

    charts = []  # (ing_id, chart, anchor_row)

    for ing_idx, (ing_id, ing_name) in enumerate(INGREDIENTS.items()):
        base_row = 3 + ing_idx * BLOCK_SIZE  # 各成分ブロックの開始行

        l_row  = PARAM_ROWS["L"][ing_id]
        k_row  = PARAM_ROWS["k"][ing_id]
        x0_row = PARAM_ROWS["x0"][ing_id]
        actual = pen_data.get(ing_id, {})

        # ── セクションヘッダー ─────────────────────────────────────
        hdr_color = C_ING_HDR[ing_idx]
        ws.merge_cells(f"A{base_row}:{get_column_letter(COL_NOTES)}{base_row}")
        c = ws.cell(row=base_row, column=1,
                    value=f"【 {ing_id}  {ing_name} 】  "
                          f"L=PARAM_L_{ing_id}  k=PARAM_k_{ing_id}  x0=PARAM_x0_{ing_id}")
        c.font      = _font(bold=True, color="FFFFFF")
        c.fill      = _fill(hdr_color)
        c.alignment = _align("left")
        ws.row_dimensions[base_row].height = 20

        # ── 列ヘッダー ─────────────────────────────────────────────
        col_headers = [
            "period", "month_index", "actual_penetration",
            "logistic_fitted（base）", "forecast_excluded",
            "forecast_adjusted（制度変更反映）", "備考",
        ]
        hdr_row = base_row + 1
        for ci, h in enumerate(col_headers, start=1):
            c = ws.cell(row=hdr_row, column=ci, value=h)
            c.font      = _font(bold=True, color="FFFFFF")
            c.fill      = _fill(C_HDR_ROW)
            c.border    = _border()
            c.alignment = _align()
        ws.row_dimensions[hdr_row].height = 18

        # ── データ行 ──────────────────────────────────────────────
        data_start = hdr_row + 1

        for i, period in enumerate(ALL_PERIODS):
            row = data_start + i
            month_idx = i
            is_forecast = i >= N_ACTUAL
            b_cell = f"B{row}"

            fill_color = C_FORECAST if is_forecast else C_ACTUAL

            # A: period
            c = ws.cell(row=row, column=COL_PERIOD, value=period)
            c.font = _font(size=10)
            c.fill = _fill(fill_color)
            c.border = _border()
            c.alignment = _align()

            # B: month_index
            c = ws.cell(row=row, column=COL_IDX, value=month_idx)
            c.font = _font(size=10)
            c.fill = _fill(fill_color)
            c.border = _border()
            c.alignment = _align()

            # C: actual_penetration（実績のみ）
            c = ws.cell(row=row, column=COL_ACTUAL)
            if not is_forecast:
                c.value = actual.get(period)
                c.number_format = "0.0%"
            c.fill = _fill(fill_color)
            c.border = _border()
            c.alignment = _align()

            # D: logistic_fitted（全期間）
            c = ws.cell(row=row, column=COL_LOGISTIC)
            c.value = _logistic_formula(b_cell, l_row, k_row, x0_row)
            c.number_format = "0.0%"
            c.fill = _fill(fill_color)
            c.border = _border()
            c.alignment = _align()

            # E: forecast_excluded（予測期間のみ; base と同じ数式）
            c = ws.cell(row=row, column=COL_FC_EXC)
            if is_forecast:
                c.value = _logistic_formula(b_cell, l_row, k_row, x0_row)
                c.number_format = "0.0%"
            c.fill = _fill(fill_color)
            c.border = _border()
            c.alignment = _align()

            # F: forecast_adjusted（予測期間のみ; 制度イベント反映）
            c = ws.cell(row=row, column=COL_FC_ADJ)
            if is_forecast:
                c.value = _logistic_adjusted_formula(b_cell, l_row, k_row, x0_row)
                c.number_format = "0.0%"
            c.fill = _fill(fill_color)
            c.border = _border()
            c.alignment = _align()

            # G: 備考
            c = ws.cell(row=row, column=COL_NOTES)
            if i == 0:
                c.value = "← 実績開始 (month_index=0)"
            elif i == N_ACTUAL:
                c.value = "← 予測開始 (month_index=41, 2026-06)"
            elif i == EVENT_START_IDX + EVENT_LAG:
                c.value = "← 制度イベント 全効果発効 (month_index=44, 2026-09)"
            c.font = _font(size=9, color="7F7F7F")
            c.fill = _fill(fill_color)
            c.border = _border()
            c.alignment = Alignment(horizontal="left")

            ws.row_dimensions[row].height = 15

        data_end = data_start + N_ALL - 1

        # ── チャート ──────────────────────────────────────────────
        chart = LineChart()
        chart.title   = f"{ing_id} {ing_name}  浸透率予測"
        chart.y_axis.title = "浸透率"
        chart.x_axis.title = "period"
        chart.y_axis.numFmt = "0%"
        chart.y_axis.scaling.min = 0.0
        chart.style = 10
        chart.width  = CHART_WIDTH
        chart.height = CHART_HEIGHT

        # series 1: actual_penetration（実績）
        ref_actual = Reference(ws, min_col=COL_ACTUAL, min_row=hdr_row,
                               max_row=data_start + N_ACTUAL - 1)
        chart.add_data(ref_actual, titles_from_data=True)

        # series 2: logistic_fitted（全期間）
        ref_fitted = Reference(ws, min_col=COL_LOGISTIC, min_row=hdr_row, max_row=data_end)
        chart.add_data(ref_fitted, titles_from_data=True)

        # series 3: forecast_adjusted（予測期間）
        ref_adj = Reference(ws, min_col=COL_FC_ADJ,
                            min_row=data_start + N_ACTUAL - 1,
                            max_row=data_end)
        chart.add_data(ref_adj, titles_from_data=False)
        chart.series[2].title = SeriesLabel(v="予測（制度変更加味）")

        # x軸ラベル（period）
        ref_x = Reference(ws, min_col=COL_PERIOD, min_row=data_start, max_row=data_end)
        chart.set_categories(ref_x)

        charts.append((ing_id, chart, base_row))

    # チャートをシートに埋め込む（各ブロックの右側 col H から）
    chart_col = get_column_letter(COL_NOTES + 2)  # I列
    for ing_idx, (ing_id, chart, base_row) in enumerate(charts):
        anchor = f"{chart_col}{base_row}"
        ws.add_chart(chart, anchor)

    print("  C_Penetration シート構築完了")
    print(f"  実績期間: {ACTUAL_PERIODS[0]} 〜 {ACTUAL_PERIODS[-1]}（{N_ACTUAL}ヶ月）")
    print(f"  予測期間: {FORECAST_PERIODS[0]} 〜 {FORECAST_PERIODS[-1]}（{N_FORECAST}ヶ月）")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# メイン
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def main():
    print(f"📂 対象ファイル: {XLSX}")
    wb = load_workbook(XLSX)

    print("① PARAM: 浸透率パラメータ初期値を更新...")
    update_param(wb)

    print("② 名前付き範囲を定義...")
    define_named_ranges(wb)

    print("③ C_Penetration シートを構築...")
    build_c_penetration(wb)

    wb.save(XLSX)
    print(f"\n✅ 保存完了: {XLSX}")


if __name__ == "__main__":
    main()
