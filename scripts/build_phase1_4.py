"""
Phase 1-4: C_ShareForecast（Step3）+ C_SKUMix（Step4）の実装
"""
import csv
from pathlib import Path
from collections import defaultdict

from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "docs" / "sample_data"
XLSX     = BASE_DIR / "output" / "GE_BS_Forecast_Engine.xlsx"

# ── スタイル ──────────────────────────────────────────────────────────
def _font(bold=False, size=11, color="000000", name="Yu Gothic"):
    return Font(bold=bold, size=size, color=color, name=name)
def _fill(h): return PatternFill("solid", fgColor=h)
def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)
def _align(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _cell(ws, row, col, value=None, fill=None, bold=False, size=11,
          color="000000", fmt=None, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    if fill: c.fill = _fill(fill)
    c.font   = _font(bold=bold, size=size, color=color)
    c.border = _border()
    c.alignment = _align(align)
    if fmt: c.number_format = fmt
    return c

# ── 定数 ─────────────────────────────────────────────────────────────
INGREDIENTS = {
    "ING01": "メトホルミン塩酸塩",
    "ING02": "アトルバスタチン",
    "ING03": "フィルグラスチム",
    "ING04": "インフリキシマブ",
}
ING_HDR_COLORS = {"ING01":"1F4E79","ING02":"375623","ING03":"4A235A","ING04":"7B241C"}
PROD_HDR_COLORS = {"PROD01":"1F4E79","PROD02":"375623","PROD03":"4A235A","PROD04":"7B241C"}

PRODUCTS = {"PROD01":"PROD01（ING01）","PROD02":"PROD02（ING02）",
            "PROD03":"PROD03（ING03）","PROD04":"PROD04（ING04）"}
PROD_SKUS = {
    "PROD01": ["SKU0101","SKU0102"],
    "PROD02": ["SKU0201","SKU0202"],
    "PROD03": ["SKU0301","SKU0302"],
    "PROD04": ["SKU0401"],
}
PROD_ING = {"PROD01":"ING01","PROD02":"ING02","PROD03":"ING03","PROD04":"ING04"}

# PARAMシートの競合パラメータ行（列E）
# 確認済み: 新規競合参入時期(ING01)=E28, 想定競合シェア(ING01)=E29
#           新規競合参入時期(ING02)=E30, 想定競合シェア(ING02)=E31
COMP_PARAM_ROWS = {
    "ING01": {"entry": 28, "share": 29},
    "ING02": {"entry": 30, "share": 31},
    "ING03": {"entry": None, "share": None},
    "ING04": {"entry": None, "share": None},
}
# PARAM SKUトレンド係数行（列E）
SKU_TREND_ROWS = {
    "SKU0101":39,"SKU0102":40,"SKU0201":41,"SKU0202":42,
    "SKU0301":43,"SKU0302":44,"SKU0401":45,
}

N_ACT = 41   # 2023-01〜2026-05
N_FC  = 36   # 2026-06〜2029-05

def _gen_periods(start_ym, n):
    y, m = int(start_ym[:4]), int(start_ym[5:7])
    out = []
    for _ in range(n):
        out.append(f"{y:04d}-{m:02d}")
        m += 1
        if m > 12: m, y = 1, y + 1
    return out

ACT_PERIODS = _gen_periods("2023-01", N_ACT)
FC_PERIODS  = _gen_periods("2026-06", N_FC)

C_ACT = "EBF5FB"
C_FC  = "FEF9E7"
C_HDR = "1A5276"

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# データロード
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _load_iqvia_share():
    """成分別・月次の 自社units / GE・BS全体units を返す"""
    with open(DATA_DIR / "master_ingredients.csv", encoding="utf-8-sig") as f:
        name_to_id = {r["name"]: r["ingredient_id"] for r in csv.DictReader(f)}

    ge_total = defaultdict(lambda: defaultdict(float))
    own      = defaultdict(lambda: defaultdict(float))
    with open(DATA_DIR / "iqvia_market_data.csv", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            ing_id = name_to_id.get(r["ingredient_name"])
            if not ing_id: continue
            p = r["period"]; u = float(r["sales_units"])
            mt = r["manufacturer_type"]
            if mt in ("自社","競合A","競合B"):
                ge_total[ing_id][p] += u
            if mt == "自社":
                own[ing_id][p] += u

    result = {}
    for ing_id in INGREDIENTS:
        result[ing_id] = {}
        for p in ACT_PERIODS:
            t = ge_total[ing_id].get(p, 0)
            o = own[ing_id].get(p, 0)
            result[ing_id][p] = {
                "own_units":      o,
                "ge_total_units": t,
                "own_share":      o / t if t else None,
            }
    return result


def _load_sellout_sku():
    """製品別・SKU別・月次の販売数量を返す"""
    with open(DATA_DIR / "master_skus.csv", encoding="utf-8-sig") as f:
        sku_to_prod = {r["sku_id"]: r["product_id"] for r in csv.DictReader(f)}

    prod_total = defaultdict(lambda: defaultdict(float))
    sku_qty    = defaultdict(lambda: defaultdict(float))
    with open(DATA_DIR / "sellout_data.csv", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            sku_id = r["sku_id"]; p = r["period"]; qty = float(r["quantity"])
            prod = sku_to_prod.get(sku_id,"")
            prod_total[prod][p] += qty
            sku_qty[sku_id][p]  += qty

    return prod_total, sku_qty


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# C_ShareForecast（Step3）
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 列レイアウト:
# A: period  B: month_index  C: own_units  D: ge_total_units
# E: own_share_actual  F: share_trend（全行）
# G: fc_share_excluded（予測のみ）  H: fc_share_adjusted（予測のみ、競合反映）
# I: notes

def _comp_entry_idx_formula(ing_id):
    """PARAM entry_date(文字列 "YYYY-MM")をmonth_indexに変換するExcel式"""
    ep = COMP_PARAM_ROWS[ing_id].get("entry")
    if ep is None:
        return None
    ref = f"PARAM!$E${ep}"
    # (YYYY - 2023)*12 + MM - 1
    return f"(VALUE(LEFT({ref},4))-2023)*12+VALUE(MID({ref},6,2))-1"


def build_c_share_forecast(wb, share_data):
    ws = wb["C_ShareForecast"]
    ws.delete_rows(1, ws.max_row)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    # 列幅
    widths = {1:10,2:12,3:14,4:14,5:14,6:16,7:18,8:22,9:32}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # タイトル
    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = ("C_ShareForecast  ―  Step3: 自社製品シェア予測"
               "（TREND + 競合参入による段階的低下）")
    c.font = _font(bold=True, size=12, color="FFFFFF")
    c.fill = _fill("4A235A"); c.alignment = _align()
    ws.row_dimensions[1].height = 24

    BLOCK = 1 + 1 + N_ACT + N_FC + 2   # hdr + col_hdr + actual + forecast + gap
    charts = []

    for ing_idx, (ing_id, ing_name) in enumerate(INGREDIENTS.items()):
        base_row  = 2 + ing_idx * BLOCK
        hdr_color = ING_HDR_COLORS[ing_id]
        sd        = share_data[ing_id]
        cp        = COMP_PARAM_ROWS[ing_id]

        # セクションヘッダー
        ws.merge_cells(f"A{base_row}:I{base_row}")
        c = ws.cell(row=base_row, column=1,
                    value=f"【 {ing_id}  {ing_name} 】")
        c.font=_font(bold=True,color="FFFFFF"); c.fill=_fill(hdr_color)
        c.alignment=_align("left"); ws.row_dimensions[base_row].height=20

        # 列ヘッダー
        col_hdrs = ["period","month_index","own_units（自社）",
                    "ge_total_units","own_share（実績）",
                    "share_trend（全期間）","fc_share_excluded",
                    "fc_share_adjusted（競合反映）","備考"]
        hr = base_row + 1
        for ci, h in enumerate(col_hdrs, 1):
            c = ws.cell(row=hr, column=ci, value=h)
            c.font=_font(bold=True,color="FFFFFF"); c.fill=_fill(C_HDR)
            c.border=_border(); c.alignment=_align()
        ws.row_dimensions[hr].height = 18

        ds = hr + 1   # data start

        # 実績行
        for i, period in enumerate(ACT_PERIODS):
            row  = ds + i
            d    = sd.get(period, {})
            fill = C_ACT

            _cell(ws, row, 1, period,             fill=fill)
            _cell(ws, row, 2, i,                  fill=fill, fmt="#,##0")
            _cell(ws, row, 3, d.get("own_units"),     fill=fill, fmt="#,##0")
            _cell(ws, row, 4, d.get("ge_total_units"),fill=fill, fmt="#,##0")
            _cell(ws, row, 5, d.get("own_share"),     fill=fill, fmt="0.0%")

            # F: TREND（実績行）
            y_rng = f"$E${ds}:$E${ds+N_ACT-1}"
            x_rng = f"$B${ds}:$B${ds+N_ACT-1}"
            c = ws.cell(row=row, column=6,
                        value=f"=TREND({y_rng},{x_rng},B{row})")
            c.fill=_fill(fill); c.border=_border(); c.number_format="0.0%"; c.alignment=_align()

            # G, H: 実績行は空欄
            _cell(ws, row, 7, fill=fill)
            _cell(ws, row, 8, fill=fill)
            _cell(ws, row, 9, fill=fill)
            if i == 0:
                ws.cell(row=row, column=9).value = "← 実績開始"
                ws.cell(row=row, column=9).font  = _font(size=9, color="7F7F7F")
            ws.row_dimensions[row].height = 15

        y_rng = f"$E${ds}:$E${ds+N_ACT-1}"
        x_rng = f"$B${ds}:$B${ds+N_ACT-1}"

        # 予測行
        for j, period in enumerate(FC_PERIODS):
            row  = ds + N_ACT + j
            fill = C_FC

            _cell(ws, row, 1, period, fill=fill)
            _cell(ws, row, 2, N_ACT + j, fill=fill, fmt="#,##0")
            _cell(ws, row, 3, fill=fill)
            _cell(ws, row, 4, fill=fill)
            _cell(ws, row, 5, fill=fill)

            # F: TREND
            c = ws.cell(row=row, column=6,
                        value=f"=TREND({y_rng},{x_rng},B{row})")
            c.fill=_fill(fill); c.border=_border(); c.number_format="0.0%"; c.alignment=_align()

            # G: fc_share_excluded = TREND（制度変更非加味 = 競合なし）
            c = ws.cell(row=row, column=7,
                        value=f"=TREND({y_rng},{x_rng},B{row})")
            c.fill=_fill(fill); c.border=_border(); c.number_format="0.0%"; c.alignment=_align()

            # H: fc_share_adjusted（競合参入反映）
            ep  = cp.get("entry")
            sp  = cp.get("share")
            if ep and sp:
                entry_ref = f"PARAM!$E${ep}"
                share_ref = f"PARAM!$E${sp}"
                # entry_idx = (YYYY-2023)*12 + MM - 1
                entry_idx = f"((VALUE(LEFT({entry_ref},4))-2023)*12+VALUE(MID({entry_ref},6,2))-1)"
                ramp      = f"MIN(MAX((B{row}-{entry_idx})/3,0),1)"
                trend_val = f"TREND({y_rng},{x_rng},B{row})"
                adj_val   = f"MAX({trend_val}-{share_ref}*{ramp},0)"
                formula   = f"=IF({entry_ref}=\"\",{trend_val},{adj_val})"
            else:
                # 競合パラメータなし → TRENDと同値
                formula = f"=TREND({y_rng},{x_rng},B{row})"

            c = ws.cell(row=row, column=8, value=formula)
            c.fill=_fill(fill); c.border=_border(); c.number_format="0.0%"; c.alignment=_align()

            c = ws.cell(row=row, column=9)
            if j == 0:
                c.value = "← 予測開始"; c.font = _font(size=9, color="7F7F7F")
            c.fill=_fill(fill); c.border=_border()
            ws.row_dimensions[row].height = 15

        data_end = ds + N_ACT + N_FC - 1

        # グラフ（自社シェア実績 + TREND + 調整済み）
        chart = LineChart()
        chart.title = f"{ing_id} {ing_name}  自社シェア予測"
        chart.y_axis.title = "自社シェア（GE/BS全体比）"
        chart.y_axis.numFmt = "0%"; chart.style = 10
        chart.width = 20; chart.height = 12

        chart.add_data(Reference(ws, min_col=5, min_row=hr, max_row=ds+N_ACT-1), titles_from_data=True)
        chart.series[0].title = SeriesLabel(v="実績シェア")

        chart.add_data(Reference(ws, min_col=6, min_row=hr, max_row=data_end), titles_from_data=True)
        chart.series[1].title = SeriesLabel(v="TRENDライン")

        chart.add_data(Reference(ws, min_col=8, min_row=ds+N_ACT-1, max_row=data_end), titles_from_data=False)
        chart.series[2].title = SeriesLabel(v="予測（競合反映）")

        ref_x = Reference(ws, min_col=1, min_row=ds, max_row=data_end)
        chart.set_categories(ref_x)
        charts.append((chart, base_row))

    chart_col = get_column_letter(10 + 2)
    for chart, br in charts:
        ws.add_chart(chart, f"{chart_col}{br}")

    print("  C_ShareForecast 構築完了")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# C_SKUMix（Step4）
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 列レイアウト（最大2 SKUの製品を想定）:
# A: period  B: month_index
# C: SKU1 qty_actual  D: SKU2 qty_actual
# E: prod_total_qty
# F: SKU1 ratio_actual  G: SKU2 ratio_actual
# H: SKU1 fc_trend_raw  I: SKU2 fc_trend_raw
# J: SKU1 fc_normalized  K: SKU2 fc_normalized
# L: notes

def build_c_sku_mix(wb, prod_total, sku_qty):
    ws = wb["C_SKUMix"]
    ws.delete_rows(1, ws.max_row)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    widths = {1:10,2:12,3:14,4:14,5:14,6:14,7:14,8:16,9:16,10:16,11:16,12:30}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.merge_cells("A1:L1")
    c = ws["A1"]
    c.value = ("C_SKUMix  ―  Step4: SKU構成比トレンド予測"
               "（TREND + MAX/MIN クリッピング + 製品内正規化）")
    c.font=_font(bold=True,size=12,color="FFFFFF"); c.fill=_fill("4A235A")
    c.alignment=_align(); ws.row_dimensions[1].height=24

    BLOCK = 1 + 1 + N_ACT + N_FC + 2
    charts = []

    for prod_idx, (prod_id, prod_name) in enumerate(PRODUCTS.items()):
        base_row  = 2 + prod_idx * BLOCK
        hdr_color = PROD_HDR_COLORS[prod_id]
        skus      = PROD_SKUS[prod_id]
        n_sku     = len(skus)
        sku1      = skus[0]
        sku2      = skus[1] if n_sku > 1 else None

        # セクションヘッダー
        ws.merge_cells(f"A{base_row}:L{base_row}")
        c = ws.cell(row=base_row, column=1,
                    value=f"【 {prod_id}  SKU: {', '.join(skus)} 】")
        c.font=_font(bold=True,color="FFFFFF"); c.fill=_fill(hdr_color)
        c.alignment=_align("left"); ws.row_dimensions[base_row].height=20

        # 列ヘッダー
        s2_label = sku2 if sku2 else "（なし）"
        col_hdrs = [
            "period","month_index",
            f"{sku1} qty",f"{s2_label} qty","prod_total",
            f"{sku1} ratio（実績）",f"{s2_label} ratio（実績）",
            f"{sku1} TREND_raw",f"{s2_label} TREND_raw",
            f"{sku1} fc_normalized",f"{s2_label} fc_normalized",
            "備考",
        ]
        hr = base_row + 1
        for ci, h in enumerate(col_hdrs, 1):
            c = ws.cell(row=hr, column=ci, value=h)
            c.font=_font(bold=True,color="FFFFFF"); c.fill=_fill(C_HDR)
            c.border=_border(); c.alignment=_align()
        ws.row_dimensions[hr].height = 18

        ds = hr + 1

        # 実績行
        for i, period in enumerate(ACT_PERIODS):
            row  = ds + i
            fill = C_ACT
            q1   = sku_qty[sku1].get(period, None)
            q2   = sku_qty[sku2].get(period, None) if sku2 else None
            qt   = prod_total[prod_id].get(period, None)

            _cell(ws, row, 1, period,      fill=fill)
            _cell(ws, row, 2, i,           fill=fill, fmt="#,##0")
            _cell(ws, row, 3, q1,          fill=fill, fmt="#,##0")
            _cell(ws, row, 4, q2,          fill=fill, fmt="#,##0")
            _cell(ws, row, 5, qt,          fill=fill, fmt="#,##0")

            # SKU1 ratio_actual
            if qt:
                _cell(ws, row, 6, f"=C{row}/E{row}", fill=fill, fmt="0.0%")
            else:
                _cell(ws, row, 6, fill=fill)

            # SKU2 ratio_actual
            if sku2 and qt:
                _cell(ws, row, 7, f"=D{row}/E{row}", fill=fill, fmt="0.0%")
            else:
                _cell(ws, row, 7, fill=fill)

            # H, I, J, K: 実績行は空欄
            for ci in [8, 9, 10, 11]:
                _cell(ws, row, ci, fill=fill)

            _cell(ws, row, 12, fill=fill)
            if i == 0:
                ws.cell(row=row, column=12).value = "← 実績開始"
                ws.cell(row=row, column=12).font  = _font(size=9, color="7F7F7F")
            ws.row_dimensions[row].height = 15

        # TREND の参照範囲
        y1_rng = f"$F${ds}:$F${ds+N_ACT-1}"
        y2_rng = f"$G${ds}:$G${ds+N_ACT-1}"
        x_rng  = f"$B${ds}:$B${ds+N_ACT-1}"

        # 予測行
        for j, period in enumerate(FC_PERIODS):
            row  = ds + N_ACT + j
            fill = C_FC

            _cell(ws, row, 1, period,       fill=fill)
            _cell(ws, row, 2, N_ACT + j,    fill=fill, fmt="#,##0")
            for ci in [3, 4, 5]:
                _cell(ws, row, ci, fill=fill)

            # F, G: 予測行は実績なし
            _cell(ws, row, 6, fill=fill)
            _cell(ws, row, 7, fill=fill)

            # H: SKU1 TREND_raw
            c = ws.cell(row=row, column=8,
                        value=f"=TREND({y1_rng},{x_rng},B{row})")
            c.fill=_fill(fill); c.border=_border(); c.number_format="0.0%"; c.alignment=_align()

            # I: SKU2 TREND_raw
            if sku2:
                c = ws.cell(row=row, column=9,
                            value=f"=TREND({y2_rng},{x_rng},B{row})")
                c.fill=_fill(fill); c.border=_border(); c.number_format="0.0%"; c.alignment=_align()
            else:
                _cell(ws, row, 9, fill=fill)

            # J: SKU1 fc_normalized
            if n_sku == 1:
                # 1 SKU 製品: 常に 1.0
                _cell(ws, row, 10, 1.0, fill=fill, fmt="0.0%")
            elif n_sku == 2:
                # MAX(0,MIN(1,H)) / (MAX(0,MIN(1,H)) + MAX(0,MIN(1,I)))
                clip1 = f"MAX(0,MIN(1,H{row}))"
                clip2 = f"MAX(0,MIN(1,I{row}))"
                total_clip = f"({clip1}+{clip2})"
                c = ws.cell(row=row, column=10,
                            value=f"=IF({total_clip}>0,{clip1}/{total_clip},0.5)")
                c.fill=_fill(fill); c.border=_border(); c.number_format="0.0%"; c.alignment=_align()

            # K: SKU2 fc_normalized
            if n_sku == 1:
                _cell(ws, row, 11, fill=fill)
            elif n_sku == 2:
                clip1 = f"MAX(0,MIN(1,H{row}))"
                clip2 = f"MAX(0,MIN(1,I{row}))"
                total_clip = f"({clip1}+{clip2})"
                c = ws.cell(row=row, column=11,
                            value=f"=IF({total_clip}>0,{clip2}/{total_clip},0.5)")
                c.fill=_fill(fill); c.border=_border(); c.number_format="0.0%"; c.alignment=_align()

            c = ws.cell(row=row, column=12)
            if j == 0:
                c.value = "← 予測開始"; c.font = _font(size=9, color="7F7F7F")
            c.fill=_fill(fill); c.border=_border()
            ws.row_dimensions[row].height = 15

        data_end = ds + N_ACT + N_FC - 1

        # グラフ（PROD01 PROD02 PROD03 のみ）—— SKU1 vs SKU2 構成比の交差を可視化
        if n_sku >= 2:
            chart = LineChart()
            chart.title = f"{prod_id} SKU 構成比トレンド（実績＋予測）"
            chart.y_axis.title = "SKU 構成比"; chart.y_axis.numFmt = "0%"
            chart.style = 10; chart.width = 20; chart.height = 12

            # 実績: SKU1 ratio
            chart.add_data(Reference(ws, min_col=6, min_row=hr, max_row=ds+N_ACT-1), titles_from_data=True)
            chart.series[0].title = SeriesLabel(v=f"{sku1} 実績")

            # 実績: SKU2 ratio
            chart.add_data(Reference(ws, min_col=7, min_row=hr, max_row=ds+N_ACT-1), titles_from_data=True)
            chart.series[1].title = SeriesLabel(v=f"{sku2} 実績")

            # 予測: SKU1 normalized（実績最終行から接続）
            chart.add_data(Reference(ws, min_col=10, min_row=ds+N_ACT-1, max_row=data_end), titles_from_data=False)
            chart.series[2].title = SeriesLabel(v=f"{sku1} 予測")

            # 予測: SKU2 normalized
            chart.add_data(Reference(ws, min_col=11, min_row=ds+N_ACT-1, max_row=data_end), titles_from_data=False)
            chart.series[3].title = SeriesLabel(v=f"{sku2} 予測")

            ref_x = Reference(ws, min_col=1, min_row=ds, max_row=data_end)
            chart.set_categories(ref_x)
            charts.append((chart, base_row))

    chart_col = get_column_letter(14)
    for chart, br in charts:
        ws.add_chart(chart, f"{chart_col}{br}")

    print("  C_SKUMix 構築完了")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PARAMシートのSKUトレンド係数セルにSLOPE数式を設定
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def update_param_sku_slopes(wb):
    """
    PARAMシートの SKUトレンド係数列に、C_SKUMixの実績ratio範囲を
    参照するSLOPE数式を書き込む（自動算出値。上書き可能）。
    """
    ws_p = wb["PARAM"]
    ws_m = wb["C_SKUMix"]

    BLOCK = 1 + 1 + N_ACT + N_FC + 2
    prod_list = list(PRODUCTS.keys())

    for prod_idx, prod_id in enumerate(prod_list):
        base_row = 2 + prod_idx * BLOCK
        hr       = base_row + 1
        ds       = hr + 1
        skus     = PROD_SKUS[prod_id]

        # SKU1: ratio_actual は 列F(=6), SKU2: 列G(=7)
        x_rng = f"C_SKUMix!$B${ds}:$B${ds+N_ACT-1}"
        for sku_col_offset, sku_id in enumerate(skus):
            ratio_col = 6 + sku_col_offset  # F=6, G=7
            ratio_rng = (f"C_SKUMix!${get_column_letter(ratio_col)}"
                         f"${ds}:${get_column_letter(ratio_col)}${ds+N_ACT-1}")
            param_row = SKU_TREND_ROWS.get(sku_id)
            if param_row:
                c = ws_p.cell(row=param_row, column=5,
                              value=f"=SLOPE({ratio_rng},{x_rng})")
                c.number_format = "0.0000"

    print("  PARAM SKUトレンド係数 SLOPE数式を設定")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def main():
    print(f"📂 {XLSX}")
    wb = load_workbook(XLSX)

    print("① データロード...")
    share_data            = _load_iqvia_share()
    prod_total, sku_qty   = _load_sellout_sku()

    print("② C_ShareForecast 構築...")
    build_c_share_forecast(wb, share_data)

    print("③ C_SKUMix 構築...")
    build_c_sku_mix(wb, prod_total, sku_qty)

    print("④ PARAM SKUトレンド係数 更新...")
    update_param_sku_slopes(wb)

    wb.save(XLSX)
    print(f"\n✅ 保存完了: {XLSX}")


if __name__ == "__main__":
    main()
