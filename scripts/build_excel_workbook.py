"""
Phase 1-1: GE_BS_Forecast_Engine.xlsx の雛形作成
  - 仕様書 5.1 節のシート構成に従い全シートを作成
  - docs/sample_data/ の CSV データをマスタ・データシートに投入
  - 各シートのデータを Excel テーブル（ListObject）として定義
  - PARAM シートに 5.2 節のパラメータ一覧をセクション分けして配置
"""
import csv
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ── パス ──────────────────────────────────────────────────────────────
BASE_DIR   = Path(__file__).resolve().parent.parent
DATA_DIR   = BASE_DIR / "docs" / "sample_data"
OUTPUT_DIR = BASE_DIR / "output"
OUTPUT_DIR.mkdir(exist_ok=True)
OUT_FILE   = OUTPUT_DIR / "GE_BS_Forecast_Engine.xlsx"

# ── カラーパレット ─────────────────────────────────────────────────────
C_HEADER_DARK  = "1F4E79"   # 濃い青（マスタ系ヘッダー）
C_HEADER_GREEN = "375623"   # 濃い緑（データ系ヘッダー）
C_HEADER_CALC  = "4A235A"   # 濃い紫（計算シート系ヘッダー）
C_HEADER_OUT   = "7B241C"   # 濃い赤（出力シート系ヘッダー）
C_HEADER_PARAM = "1A5276"   # PARAMヘッダー
C_SECTION      = "D6E4F0"   # セクション行背景（薄い青）
C_SECTION_GRN  = "D5F5E3"   # セクション行背景（薄い緑）
C_SECTION_YLW  = "FDFEFE"   # 薄いグレー
C_WHITE        = "FFFFFF"
C_LIGHT_GRAY   = "F2F2F2"

# ── スタイルヘルパー ───────────────────────────────────────────────────
def _font(bold=False, size=11, color="000000", name="Yu Gothic"):
    return Font(bold=bold, size=size, color=color, name=name)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border_thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _header_cell(ws, row, col, value, bg_color, font_color="FFFFFF"):
    c = ws.cell(row=row, column=col, value=value)
    c.font    = _font(bold=True, color=font_color)
    c.fill    = _fill(bg_color)
    c.border  = _border_thin()
    c.alignment = _align("center")
    return c

def _add_table(ws, min_row, min_col, max_row, max_col, tbl_name, style="TableStyleMedium9"):
    ref = (
        f"{get_column_letter(min_col)}{min_row}:"
        f"{get_column_letter(max_col)}{max_row}"
    )
    tbl = Table(displayName=tbl_name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name=style, showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False,
    )
    ws.add_table(tbl)

def _autofit(ws, min_col=1, max_col=None, padding=4):
    """列幅を内容に合わせて自動調整（openpyxl は自動計算不可のため文字数で推算）"""
    max_col = max_col or ws.max_column
    for col_idx in range(min_col, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + padding, 40)

def _read_csv(filename):
    path = DATA_DIR / filename
    with open(path, encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        rows = list(reader)
    return rows  # rows[0] = header, rows[1:] = data

# ── シート名リスト（仕様書 5.1 節の順序） ─────────────────────────────
SHEET_ORDER = [
    "START",
    "PARAM",
    "M_TherapeuticArea",
    "M_Ingredient",
    "M_Product",
    "M_SKU",
    "M_Events",
    "M_FXRate",
    "D_IQVIA",
    "D_SellIn",
    "D_SellOut",
    "D_Inventory",
    "C_MarketForecast",
    "C_Penetration",
    "C_ShareForecast",
    "C_SKUMix",
    "C_Integrated",
    "OUT_Adjusted",
    "OUT_Excluded",
    "LOG",
]

# ── シートタブカラー ──────────────────────────────────────────────────
SHEET_COLORS = {
    "START":              "808080",
    "PARAM":              "2E86C1",
    "M_TherapeuticArea":  "1F4E79",
    "M_Ingredient":       "1F4E79",
    "M_Product":          "1F4E79",
    "M_SKU":              "1F4E79",
    "M_Events":           "1F4E79",
    "M_FXRate":           "1F4E79",
    "D_IQVIA":            "375623",
    "D_SellIn":           "375623",
    "D_SellOut":          "375623",
    "D_Inventory":        "375623",
    "C_MarketForecast":   "4A235A",
    "C_Penetration":      "4A235A",
    "C_ShareForecast":    "4A235A",
    "C_SKUMix":           "4A235A",
    "C_Integrated":       "4A235A",
    "OUT_Adjusted":       "7B241C",
    "OUT_Excluded":       "7B241C",
    "LOG":                "5D5D5D",
}

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# START シート
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _build_start(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 60

    # タイトル
    ws.merge_cells("B1:C1")
    c = ws["B1"]
    c.value = "GE/BS 販売予測エンジン（Excel版）"
    c.font  = _font(bold=True, size=16, color="1F4E79")
    c.alignment = _align("center")

    ws.merge_cells("B2:C2")
    c = ws["B2"]
    c.value = "GE_BS_Forecast_Engine.xlsx  |  Phase 1-1 初期版"
    c.font  = _font(size=10, color="7F7F7F")
    c.alignment = _align("center")

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18

    # シート一覧
    rows = [
        (4,  "シート名",          "内容",                                         True),
        (5,  "START",            "本ガイドシート（読み取り専用）",                  False),
        (6,  "PARAM",            "全パラメータ一覧（ユーザー編集）",                False),
        (7,  "M_TherapeuticArea","疾患領域マスタ",                                 False),
        (8,  "M_Ingredient",     "成分マスタ",                                    False),
        (9,  "M_Product",        "製品マスタ",                                    False),
        (10, "M_SKU",            "SKUマスタ",                                     False),
        (11, "M_Events",         "制度イベントマスタ",                              False),
        (12, "M_FXRate",         "為替レートマスタ",                               False),
        (13, "D_IQVIA",          "IQVIA市場データ（Power Query取込）",              False),
        (14, "D_SellIn",         "自社Sell-inデータ",                             False),
        (15, "D_SellOut",        "自社Sell-outデータ",                            False),
        (16, "D_Inventory",      "卸別月末在庫データ",                             False),
        (17, "C_MarketForecast", "Step1: 市場トレンド予測（FORECAST.ETS）",        False),
        (18, "C_Penetration",    "Step2: 浸透率予測（ロジスティック曲線）",         False),
        (19, "C_ShareForecast",  "Step3: 自社製品シェア予測",                     False),
        (20, "C_SKUMix",         "Step4: SKU構成比トレンド予測",                  False),
        (21, "C_Integrated",     "Step5: 統合予測（シナリオ別）",                  False),
        (22, "OUT_Adjusted",     "出力: 制度変更加味シナリオ",                     False),
        (23, "OUT_Excluded",     "出力: 制度変更非加味シナリオ",                   False),
        (24, "LOG",              "パラメータ変更履歴（VBA自動記録）",               False),
    ]
    for r, name, desc, is_hdr in rows:
        if is_hdr:
            _header_cell(ws, r, 2, name, C_HEADER_PARAM)
            _header_cell(ws, r, 3, desc, C_HEADER_PARAM)
        else:
            c1 = ws.cell(row=r, column=2, value=name)
            c2 = ws.cell(row=r, column=3, value=desc)
            if r % 2 == 0:
                c1.fill = c2.fill = _fill(C_LIGHT_GRAY)
            for c in [c1, c2]:
                c.border = _border_thin()
                c.alignment = _align()

    # 月次更新手順
    ws.row_dimensions[26].height = 16
    t = ws.cell(row=26, column=2, value="■ 月次更新手順")
    t.font = _font(bold=True, size=12, color="1F4E79")
    steps = [
        "① CSVデータを docs/sample_data/ に配置",
        "② [データ] → [すべて更新] で Power Query を実行",
        "③ PARAM シートでパラメータを確認・調整",
        "④ Ctrl+Alt+F9 でブック全体を再計算",
        "⑤ Power BI (.pbix) を開いて [更新] をクリック",
    ]
    for i, s in enumerate(steps):
        c = ws.cell(row=27 + i, column=2, value=s)
        c.font = _font(size=10)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PARAM シート（仕様書 5.2 節）
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _build_param(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 22   # セクション
    ws.column_dimensions["C"].width = 32   # パラメータ名
    ws.column_dimensions["D"].width = 18   # 対象（成分/SKU等）
    ws.column_dimensions["E"].width = 16   # 値（ユーザー入力）
    ws.column_dimensions["F"].width = 14   # 単位
    ws.column_dimensions["G"].width = 36   # 備考

    # タイトル
    ws.merge_cells("B1:G1")
    c = ws["B1"]
    c.value     = "PARAM  ―  パラメータ管理シート"
    c.font      = _font(bold=True, size=14, color="FFFFFF")
    c.fill      = _fill(C_HEADER_PARAM)
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 28

    # 列ヘッダー
    headers = ["セクション", "パラメータ", "対象", "値（入力）", "単位", "備考"]
    for i, h in enumerate(headers, start=2):
        _header_cell(ws, 2, i, h, C_HEADER_PARAM)
    ws.row_dimensions[2].height = 20

    # パラメータ定義
    # (section, param, target, default_value, unit, note)
    PARAMS = [
        # ── 薬価セクション ──────────────────────────────────────────
        ("薬価", "次回改定予定月",         "",       "2026-04", "年-月", "次回改定の発効月"),
        ("薬価", "改定率（ING01）",        "ING01",  -0.055,   "%",    "過去実績平均値"),
        ("薬価", "改定率（ING02）",        "ING02",  -0.062,   "%",    "過去実績平均値"),
        ("薬価", "改定率（ING03）",        "ING03",  -0.048,   "%",    "過去実績平均値"),
        ("薬価", "改定率（ING04）",        "ING04",  -0.070,   "%",    "過去実績平均値"),
        # ── 浸透率セクション ────────────────────────────────────────
        ("浸透率", "浸透率上限L（ING01）", "ING01",  0.95,     "0〜1", "Solverフィッティング結果"),
        ("浸透率", "浸透率上限L（ING02）", "ING02",  0.90,     "0〜1", "Solverフィッティング結果"),
        ("浸透率", "浸透率上限L（ING03）", "ING03",  0.575,    "0〜1", "Solverフィッティング結果"),
        ("浸透率", "浸透率上限L（ING04）", "ING04",  0.430,    "0〜1", "Solverフィッティング結果"),
        ("浸透率", "立ち上がり速度k（ING01）", "ING01", 0.0512, "正の実数", "Solverフィッティング結果"),
        ("浸透率", "立ち上がり速度k（ING02）", "ING02", 0.0489, "正の実数", "Solverフィッティング結果"),
        ("浸透率", "立ち上がり速度k（ING03）", "ING03", 0.1823, "正の実数", "Solverフィッティング結果"),
        ("浸透率", "立ち上がり速度k（ING04）", "ING04", 0.2104, "正の実数", "Solverフィッティング結果"),
        ("浸透率", "変曲点x0（ING01）",    "ING01",  18.5,    "月数", "実績開始からの経過月"),
        ("浸透率", "変曲点x0（ING02）",    "ING02",  22.1,    "月数", "実績開始からの経過月"),
        ("浸透率", "変曲点x0（ING03）",    "ING03",  14.3,    "月数", "実績開始からの経過月"),
        ("浸透率", "変曲点x0（ING04）",    "ING04",  12.8,    "月数", "実績開始からの経過月"),
        # ── 市場セクション ──────────────────────────────────────────
        ("市場", "疾患領域市場成長率（TA01）", "TA01", 0.015, "%/年", "代謝・内分泌領域"),
        ("市場", "疾患領域市場成長率（TA02）", "TA02", 0.010, "%/年", "循環器領域"),
        ("市場", "疾患領域市場成長率（TA03）", "TA03", 0.008, "%/年", "血液領域"),
        ("市場", "疾患領域市場成長率（TA04）", "TA04", 0.025, "%/年", "免疫領域"),
        # ── 競合セクション ──────────────────────────────────────────
        ("競合", "新規競合参入時期（ING01）", "ING01", "2026-10", "年-月", "手動入力"),
        ("競合", "想定競合シェア（ING01）",  "ING01", 0.05,  "%",    "参入後の最大シェア"),
        ("競合", "新規競合参入時期（ING02）", "ING02", "",    "年-月", "未定の場合は空欄"),
        ("競合", "想定競合シェア（ING02）",  "ING02", "",    "%",    "未定の場合は空欄"),
        # ── 在庫セクション ──────────────────────────────────────────
        ("在庫", "適正在庫日数（卸A）",     "卸A",   30,   "日",   "実測平均"),
        ("在庫", "適正在庫日数（卸B）",     "卸B",   28,   "日",   "実測平均"),
        ("在庫", "適正在庫日数（卸C）",     "卸C",   32,   "日",   "実測平均"),
        ("在庫", "改定前積み増し率",        "",       0.15, "%",    "薬価改定前の卸積み増し率"),
        ("在庫", "改定後調整率",            "",      -0.10, "%",    "改定後の在庫取り崩し率"),
        # ── SKUミックスセクション ───────────────────────────────────
        ("SKUミックス", "SKU0101 トレンド係数", "SKU0101", "",  "自動算出", "TREND関数の結果を上書き可"),
        ("SKUミックス", "SKU0102 トレンド係数", "SKU0102", "",  "自動算出", "TREND関数の結果を上書き可"),
        ("SKUミックス", "SKU0201 トレンド係数", "SKU0201", "",  "自動算出", "TREND関数の結果を上書き可"),
        ("SKUミックス", "SKU0202 トレンド係数", "SKU0202", "",  "自動算出", "TREND関数の結果を上書き可"),
        ("SKUミックス", "SKU0301 トレンド係数", "SKU0301", "",  "自動算出", "TREND関数の結果を上書き可"),
        ("SKUミックス", "SKU0302 トレンド係数", "SKU0302", "",  "自動算出", "TREND関数の結果を上書き可"),
        ("SKUミックス", "SKU0401 トレンド係数", "SKU0401", "",  "自動算出", "TREND関数の結果を上書き可"),
        # ── 通貨セクション ──────────────────────────────────────────
        ("通貨", "予測仮定レート（JPY/USD）", "", 150.0, "円/USD", "直近実績レートを初期値とする"),
        # ── シナリオセクション ──────────────────────────────────────
        ("シナリオ", "楽観シナリオ 調整幅",  "", 0.10, "%", "ベースから+10%を初期値"),
        ("シナリオ", "悲観シナリオ 調整幅",  "", -0.10, "%", "ベースから-10%を初期値"),
    ]

    current_section = None
    row = 3
    section_colors = {
        "薬価":       "FEF9E7",
        "浸透率":     "EBF5FB",
        "市場":       "E9F7EF",
        "競合":       "FDEDEC",
        "在庫":       "F5EEF8",
        "SKUミックス":"FDF2E9",
        "通貨":       "E8F8F5",
        "シナリオ":   "F0F3F4",
    }

    for section, param, target, value, unit, note in PARAMS:
        # セクション見出し行
        if section != current_section:
            ws.merge_cells(f"B{row}:G{row}")
            c = ws.cell(row=row, column=2, value=f"【 {section} 】")
            c.font      = _font(bold=True, size=11, color="1A5276")
            c.fill      = _fill("D6EAF8")
            c.alignment = _align("left")
            c.border    = _border_thin()
            ws.row_dimensions[row].height = 20
            current_section = section
            row += 1

        fill_color = section_colors.get(section, C_WHITE)
        vals = [section, param, target, value, unit, note]
        for i, v in enumerate(vals, start=2):
            c = ws.cell(row=row, column=i, value=v)
            c.border    = _border_thin()
            c.alignment = _align()
            c.fill      = _fill(fill_color)
            if i == 5:  # 値列
                c.font = _font(bold=True, color="1A5276")
        ws.row_dimensions[row].height = 17
        row += 1

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CSV → シート 汎用ロード
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _load_csv_to_sheet(ws, csv_filename, tbl_name, hdr_color, tbl_style="TableStyleMedium9"):
    rows = _read_csv(csv_filename)
    if not rows:
        return
    headers = rows[0]
    data    = rows[1:]

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    # ヘッダー行
    for col_idx, h in enumerate(headers, start=1):
        _header_cell(ws, 1, col_idx, h, hdr_color)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(h) + 4, 14)

    # データ行
    for r_idx, row in enumerate(data, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=_coerce(val))
            cell.border    = _border_thin()
            cell.alignment = _align()
            if r_idx % 2 == 0:
                cell.fill = _fill(C_LIGHT_GRAY)

    # テーブル定義
    if data:
        _add_table(ws, 1, 1, len(data) + 1, len(headers), tbl_name, tbl_style)

    _autofit(ws)


def _coerce(val: str):
    """数値に変換できれば変換、それ以外は文字列のまま返す。"""
    if val == "":
        return None
    try:
        return int(val)
    except ValueError:
        pass
    try:
        return float(val)
    except ValueError:
        pass
    return val

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 計算・出力・LOGシートの骨格
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _build_calc_skeleton(ws, title, description, hdr_color):
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = f"{title}  ―  {description}"
    c.font      = _font(bold=True, size=12, color="FFFFFF")
    c.fill      = _fill(hdr_color)
    c.alignment = _align("center")
    ws.row_dimensions[1].height = 24

    note = ws.cell(row=3, column=1,
                   value="※ このシートは Phase 1-2 以降で数式・Python in Excel を実装します。")
    note.font      = _font(size=10, color="7F7F7F")
    note.alignment = _align()


def _build_log(ws):
    ws.sheet_view.showGridLines = False
    headers = ["timestamp", "user", "sheet", "cell_address", "parameter_name",
               "old_value", "new_value"]
    for i, h in enumerate(headers, start=1):
        _header_cell(ws, 1, i, h, "5D5D5D")
    _add_table(ws, 1, 1, 1, len(headers), "tbl_LOG", "TableStyleMedium1")
    _autofit(ws, max_col=len(headers))
    note = ws.cell(row=3, column=1,
                   value="※ PARAM シートへの入力変更は VBA（Worksheet_Change）で自動記録されます。")
    note.font = _font(size=10, color="7F7F7F")


def _build_out_skeleton(ws, tbl_name, hdr_color):
    """OUT_Adjusted / OUT_Excluded の出力テーブル骨格"""
    ws.sheet_view.showGridLines = False
    headers = [
        "period", "sku_id", "ingredient_id", "product_id",
        "scenario_axis1", "scenario_axis2",
        "sellout_forecast_units", "sellout_forecast_jpy", "sellout_forecast_usd",
        "sellin_forecast_units",  "sellin_forecast_jpy",  "sellin_forecast_usd",
    ]
    for i, h in enumerate(headers, start=1):
        _header_cell(ws, 1, i, h, hdr_color)
        ws.column_dimensions[get_column_letter(i)].width = max(len(h) + 4, 16)

    # 空テーブル（1行だけ確保して ListObject を定義）
    _add_table(ws, 1, 1, 1, len(headers), tbl_name, "TableStyleMedium3")

    note = ws.cell(row=3, column=1,
                   value="※ C_Integrated シート（Phase 1-5）の数式により自動生成されます。")
    note.font = _font(size=10, color="7F7F7F")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# メイン
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def main():
    wb = Workbook()
    # デフォルトシートを削除
    wb.remove(wb.active)

    # 全シートを順番通りに作成
    for name in SHEET_ORDER:
        ws = wb.create_sheet(name)
        ws.sheet_properties.tabColor = SHEET_COLORS.get(name, "FFFFFF")

    # ── START ───────────────────────────────────────────────────────
    _build_start(wb["START"])

    # ── PARAM ───────────────────────────────────────────────────────
    _build_param(wb["PARAM"])

    # ── マスタシート（CSV → テーブル） ─────────────────────────────
    _load_csv_to_sheet(
        wb["M_TherapeuticArea"],
        "master_therapeutic_areas.csv",
        "tbl_M_TherapeuticArea",
        C_HEADER_DARK,
        "TableStyleMedium9",
    )
    _load_csv_to_sheet(
        wb["M_Ingredient"],
        "master_ingredients.csv",
        "tbl_M_Ingredient",
        C_HEADER_DARK,
    )
    _load_csv_to_sheet(
        wb["M_Product"],
        "master_products.csv",
        "tbl_M_Product",
        C_HEADER_DARK,
    )
    _load_csv_to_sheet(
        wb["M_SKU"],
        "master_skus.csv",
        "tbl_M_SKU",
        C_HEADER_DARK,
    )
    _load_csv_to_sheet(
        wb["M_Events"],
        "regulatory_events_sample.csv",
        "tbl_M_Events",
        C_HEADER_DARK,
    )
    _load_csv_to_sheet(
        wb["M_FXRate"],
        "fx_rates_sample.csv",
        "tbl_M_FXRate",
        C_HEADER_DARK,
    )

    # ── データシート（CSV → テーブル） ─────────────────────────────
    _load_csv_to_sheet(
        wb["D_IQVIA"],
        "iqvia_market_data.csv",
        "tbl_D_IQVIA",
        C_HEADER_GREEN,
        "TableStyleMedium14",
    )
    _load_csv_to_sheet(
        wb["D_SellIn"],
        "sellin_data.csv",
        "tbl_D_SellIn",
        C_HEADER_GREEN,
        "TableStyleMedium14",
    )
    _load_csv_to_sheet(
        wb["D_SellOut"],
        "sellout_data.csv",
        "tbl_D_SellOut",
        C_HEADER_GREEN,
        "TableStyleMedium14",
    )
    _load_csv_to_sheet(
        wb["D_Inventory"],
        "inventory_data.csv",
        "tbl_D_Inventory",
        C_HEADER_GREEN,
        "TableStyleMedium14",
    )

    # ── 計算シート骨格 ───────────────────────────────────────────
    _build_calc_skeleton(wb["C_MarketForecast"], "C_MarketForecast",
                         "Step1: 成分別市場トレンド予測（FORECAST.ETS）", C_HEADER_CALC)
    _build_calc_skeleton(wb["C_Penetration"],    "C_Penetration",
                         "Step2: 浸透率予測（ロジスティック曲線 / Python in Excel）", C_HEADER_CALC)
    _build_calc_skeleton(wb["C_ShareForecast"],  "C_ShareForecast",
                         "Step3: 自社製品シェア予測（TREND + 競合パラメータ）", C_HEADER_CALC)
    _build_calc_skeleton(wb["C_SKUMix"],         "C_SKUMix",
                         "Step4: SKU構成比トレンド予測（TREND + 正規化）", C_HEADER_CALC)
    _build_calc_skeleton(wb["C_Integrated"],     "C_Integrated",
                         "Step5: 統合予測（シナリオ別 SKU×月次）", C_HEADER_CALC)

    # ── 出力シート骨格 ───────────────────────────────────────────
    _build_out_skeleton(wb["OUT_Adjusted"], "tbl_OUT_Adjusted", C_HEADER_OUT)
    _build_out_skeleton(wb["OUT_Excluded"], "tbl_OUT_Excluded", C_HEADER_OUT)

    # ── LOG ─────────────────────────────────────────────────────
    _build_log(wb["LOG"])

    # 保存
    wb.save(OUT_FILE)
    print(f"✅ 出力完了: {OUT_FILE}")
    print(f"   シート数: {len(wb.sheetnames)}")
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"   {name:<22} 最終行={ws.max_row}  最終列={ws.max_column}")


if __name__ == "__main__":
    main()
