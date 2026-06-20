"""
Phase 1-2: C_MarketForecast シートの実装
  - D_IQVIAシートを参照し、成分別月次市場数量・金額の実績を集計
  - FORECAST.ETS で36ヶ月予測（季節性=12）
  - FORECAST.ETS.CONFINT で80%/95%信頼区間
  - 成分ごとに折れ線グラフを作成
"""
import csv
from pathlib import Path
from collections import defaultdict

from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "docs" / "sample_data"
XLSX     = BASE_DIR / "output" / "GE_BS_Forecast_Engine.xlsx"

# ── スタイルヘルパー ───────────────────────────────────────────────────
def _font(bold=False, size=11, color="000000", name="Yu Gothic"):
    return Font(bold=bold, size=size, color=color, name=name)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="center", v="center"):
    return Alignment(horizontal=h, vertical=v)

# ── 定数 ─────────────────────────────────────────────────────────────
INGREDIENTS = {
    "ING01": "メトホルミン塩酸塩",
    "ING02": "アトルバスタチン",
    "ING03": "フィルグラスチム",
    "ING04": "インフリキシマブ",
}
ING_HDR_COLORS = ["1F4E79", "375623", "4A235A", "7B241C"]

N_ACTUAL   = 41   # 実績: 2023-01〜2026-05
N_FORECAST = 36   # 予測: 2026-06〜2029-05
SEASONALITY = 12  # FORECAST.ETS の月次季節性

def _gen_periods(start_ym, n):
    y, m = int(start_ym[:4]), int(start_ym[5:7])
    result = []
    for _ in range(n):
        result.append(f"{y:04d}-{m:02d}")
        m += 1
        if m > 12:
            m, y = 1, y + 1
    return result

ACTUAL_PERIODS   = _gen_periods("2023-01", N_ACTUAL)
FORECAST_PERIODS = _gen_periods("2026-06", N_FORECAST)

# ── IQVIAデータを成分別に集計 ─────────────────────────────────────────
def _load_iqvia():
    with open(DATA_DIR / "master_ingredients.csv", encoding="utf-8-sig") as f:
        ing_name_to_id = {r["name"]: r["ingredient_id"] for r in csv.DictReader(f)}

    totals = defaultdict(lambda: defaultdict(lambda: {"units": 0.0, "amount": 0.0}))
    with open(DATA_DIR / "iqvia_market_data.csv", encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            ing_id = ing_name_to_id.get(r["ingredient_name"])
            if ing_id:
                totals[ing_id][r["period"]]["units"]  += float(r["sales_units"])
                totals[ing_id][r["period"]]["amount"] += float(r["sales_amount_jpy"])
    return totals

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# C_MarketForecast シートを構築
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#
# 列レイアウト（成分ごとのブロックを縦に並べる）
#
# A: period
# B: total_units_actual     （実績数量）
# C: total_amount_jpy_actual（実績金額JPY）
# D: fc_units               （FORECAST.ETS 数量予測 ← 実績行も埋める）
# E: fc_units_lower80       （FORECAST.ETS.CONFINT 80% 下限）
# F: fc_units_upper80       （FORECAST.ETS.CONFINT 80% 上限）
# G: fc_units_lower95       （FORECAST.ETS.CONFINT 95% 下限）
# H: fc_units_upper95       （FORECAST.ETS.CONFINT 95% 上限）
# I: fc_amount_jpy          （FORECAST.ETS 金額予測）
# J: notes

COL_PERIOD    = 1   # A
COL_ACT_U     = 2   # B
COL_ACT_A     = 3   # C
COL_FC_U      = 4   # D  FORECAST.ETS (units)
COL_FC_U_L80  = 5   # E
COL_FC_U_U80  = 6   # F
COL_FC_U_L95  = 7   # G
COL_FC_U_U95  = 8   # H
COL_FC_A      = 9   # I  FORECAST.ETS (amount_jpy)
COL_NOTES     = 10  # J

N_COLS = COL_NOTES

# 1ブロック = section_hdr(1) + col_hdr(1) + actual(N_ACTUAL) + forecast(N_FORECAST) + gap(2)
BLOCK = 1 + 1 + N_ACTUAL + N_FORECAST + 2

C_ACTUAL   = "EBF5FB"
C_FORECAST = "FEF9E7"
C_HDR      = "1A5276"


def build_c_market_forecast(wb):
    ws = wb["C_MarketForecast"]
    ws.delete_rows(1, ws.max_row)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    # 列幅
    widths = {1:10, 2:16, 3:18, 4:16, 5:16, 6:16, 7:16, 8:16, 9:18, 10:30}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # タイトル
    ws.merge_cells(f"A1:{get_column_letter(N_COLS)}1")
    c = ws["A1"]
    c.value     = ("C_MarketForecast  ―  Step1: 成分別市場サイズ予測"
                   "（FORECAST.ETS, 季節性=12, 信頼区間80%/95%）")
    c.font      = _font(bold=True, size=12, color="FFFFFF")
    c.fill      = _fill("4A235A")
    c.alignment = _align()
    ws.row_dimensions[1].height = 24

    iqvia = _load_iqvia()

    charts = []

    for ing_idx, (ing_id, ing_name) in enumerate(INGREDIENTS.items()):
        base_row  = 2 + ing_idx * BLOCK
        hdr_color = ING_HDR_COLORS[ing_idx]
        data_rows = iqvia.get(ing_id, {})

        # ── セクションヘッダー ──────────────────────────────────
        ws.merge_cells(f"A{base_row}:{get_column_letter(N_COLS)}{base_row}")
        c = ws.cell(row=base_row, column=1,
                    value=f"【 {ing_id}  {ing_name} 】")
        c.font      = _font(bold=True, color="FFFFFF")
        c.fill      = _fill(hdr_color)
        c.alignment = _align("left")
        ws.row_dimensions[base_row].height = 20

        # ── 列ヘッダー ────────────────────────────────────────
        col_hdrs = [
            "period", "total_units（実績）", "total_amount_jpy（実績）",
            "fc_units（ETS）", "fc_units_lower80", "fc_units_upper80",
            "fc_units_lower95", "fc_units_upper95",
            "fc_amount_jpy（ETS）", "備考",
        ]
        hdr_row = base_row + 1
        for ci, h in enumerate(col_hdrs, start=1):
            c = ws.cell(row=hdr_row, column=ci, value=h)
            c.font      = _font(bold=True, color="FFFFFF")
            c.fill      = _fill(C_HDR)
            c.border    = _border()
            c.alignment = _align()
        ws.row_dimensions[hdr_row].height = 18

        data_start = hdr_row + 1

        # ── 実績行 ────────────────────────────────────────────
        for i, period in enumerate(ACTUAL_PERIODS):
            row = data_start + i
            d   = data_rows.get(period, {"units": None, "amount": None})

            # A: period
            c = ws.cell(row=row, column=COL_PERIOD, value=period)
            c.fill = _fill(C_ACTUAL); c.border = _border(); c.alignment = _align()

            # B: 実績数量
            c = ws.cell(row=row, column=COL_ACT_U, value=d["units"])
            c.fill = _fill(C_ACTUAL); c.border = _border(); c.alignment = _align()
            c.number_format = "#,##0"

            # C: 実績金額
            c = ws.cell(row=row, column=COL_ACT_A, value=d["amount"])
            c.fill = _fill(C_ACTUAL); c.border = _border(); c.alignment = _align()
            c.number_format = "#,##0"

            # D-I: 実績行はFORECAST.ETSのフィッティング値も入れる
            # 実績行に対するフィッティングは数式では直接出ないため空欄（グラフは実績列を使用）
            for ci in range(COL_FC_U, COL_NOTES + 1):
                c = ws.cell(row=row, column=ci)
                c.fill = _fill(C_ACTUAL); c.border = _border()

            # J: 備考
            if i == 0:
                ws.cell(row=row, column=COL_NOTES, value="← 実績開始").font = _font(size=9, color="7F7F7F")
            ws.row_dimensions[row].height = 15

        # 実績範囲の参照文字列（FORECAST.ETS の known_y / known_x として使用）
        act_u_ref   = f"$B${data_start}:$B${data_start + N_ACTUAL - 1}"   # 実績数量
        act_a_ref   = f"$C${data_start}:$C${data_start + N_ACTUAL - 1}"   # 実績金額
        act_p_ref   = f"$A${data_start}:$A${data_start + N_ACTUAL - 1}"   # 実績期間（x軸）

        # ── 予測行 ────────────────────────────────────────────
        for j, period in enumerate(FORECAST_PERIODS):
            row = data_start + N_ACTUAL + j

            # A: period
            c = ws.cell(row=row, column=COL_PERIOD, value=period)
            c.fill = _fill(C_FORECAST); c.border = _border(); c.alignment = _align()

            # B: 実績なし
            c = ws.cell(row=row, column=COL_ACT_U)
            c.fill = _fill(C_FORECAST); c.border = _border()

            # C: 実績なし
            c = ws.cell(row=row, column=COL_ACT_A)
            c.fill = _fill(C_FORECAST); c.border = _border()

            a_cell = f"A{row}"  # 予測対象の期間セル

            # D: FORECAST.ETS（数量）
            # =FORECAST.ETS(target_date, values, timeline, seasonality, data_completion, aggregation)
            f_u = (f'=FORECAST.ETS({a_cell},{act_u_ref},{act_p_ref}'
                   f',{SEASONALITY},1)')
            c = ws.cell(row=row, column=COL_FC_U, value=f_u)
            c.fill = _fill(C_FORECAST); c.border = _border(); c.alignment = _align()
            c.number_format = "#,##0"

            # E: 下限80%
            f_l80 = (f'=FORECAST.ETS({a_cell},{act_u_ref},{act_p_ref}'
                     f',{SEASONALITY},1)-FORECAST.ETS.CONFINT({a_cell},{act_u_ref},{act_p_ref}'
                     f',0.8,{SEASONALITY},1)')
            c = ws.cell(row=row, column=COL_FC_U_L80, value=f_l80)
            c.fill = _fill(C_FORECAST); c.border = _border(); c.alignment = _align()
            c.number_format = "#,##0"

            # F: 上限80%
            f_u80 = (f'=FORECAST.ETS({a_cell},{act_u_ref},{act_p_ref}'
                     f',{SEASONALITY},1)+FORECAST.ETS.CONFINT({a_cell},{act_u_ref},{act_p_ref}'
                     f',0.8,{SEASONALITY},1)')
            c = ws.cell(row=row, column=COL_FC_U_U80, value=f_u80)
            c.fill = _fill(C_FORECAST); c.border = _border(); c.alignment = _align()
            c.number_format = "#,##0"

            # G: 下限95%
            f_l95 = (f'=FORECAST.ETS({a_cell},{act_u_ref},{act_p_ref}'
                     f',{SEASONALITY},1)-FORECAST.ETS.CONFINT({a_cell},{act_u_ref},{act_p_ref}'
                     f',0.95,{SEASONALITY},1)')
            c = ws.cell(row=row, column=COL_FC_U_L95, value=f_l95)
            c.fill = _fill(C_FORECAST); c.border = _border(); c.alignment = _align()
            c.number_format = "#,##0"

            # H: 上限95%
            f_u95 = (f'=FORECAST.ETS({a_cell},{act_u_ref},{act_p_ref}'
                     f',{SEASONALITY},1)+FORECAST.ETS.CONFINT({a_cell},{act_u_ref},{act_p_ref}'
                     f',0.95,{SEASONALITY},1)')
            c = ws.cell(row=row, column=COL_FC_U_U95, value=f_u95)
            c.fill = _fill(C_FORECAST); c.border = _border(); c.alignment = _align()
            c.number_format = "#,##0"

            # I: FORECAST.ETS（金額）
            f_a = (f'=FORECAST.ETS({a_cell},{act_a_ref},{act_p_ref}'
                   f',{SEASONALITY},1)')
            c = ws.cell(row=row, column=COL_FC_A, value=f_a)
            c.fill = _fill(C_FORECAST); c.border = _border(); c.alignment = _align()
            c.number_format = "#,##0"

            # J: 備考
            c = ws.cell(row=row, column=COL_NOTES)
            if j == 0:
                c.value = "← 予測開始 (2026-06)"
                c.font  = _font(size=9, color="7F7F7F")
            c.fill = _fill(C_FORECAST); c.border = _border()
            ws.row_dimensions[row].height = 15

        data_end = data_start + N_ACTUAL + N_FORECAST - 1

        # ── 折れ線グラフ（数量） ──────────────────────────────
        chart = LineChart()
        chart.title         = f"{ing_id} {ing_name}  市場数量（実績＋FORECAST.ETS予測）"
        chart.y_axis.title  = "販売数量（本）"
        chart.x_axis.title  = "period"
        chart.y_axis.numFmt = "#,##0"
        chart.style         = 10
        chart.width         = 22
        chart.height        = 14

        # 実績
        ref_act = Reference(ws, min_col=COL_ACT_U,
                            min_row=hdr_row, max_row=data_start + N_ACTUAL - 1)
        chart.add_data(ref_act, titles_from_data=True)
        chart.series[0].title = SeriesLabel(v="実績数量")

        # 予測
        ref_fc = Reference(ws, min_col=COL_FC_U,
                           min_row=data_start + N_ACTUAL - 1, max_row=data_end)
        chart.add_data(ref_fc, titles_from_data=False)
        chart.series[1].title = SeriesLabel(v="予測（ETS）")

        # 上限80%
        ref_u80 = Reference(ws, min_col=COL_FC_U_U80,
                            min_row=data_start + N_ACTUAL - 1, max_row=data_end)
        chart.add_data(ref_u80, titles_from_data=False)
        chart.series[2].title = SeriesLabel(v="上限80%CI")

        # 下限80%
        ref_l80 = Reference(ws, min_col=COL_FC_U_L80,
                            min_row=data_start + N_ACTUAL - 1, max_row=data_end)
        chart.add_data(ref_l80, titles_from_data=False)
        chart.series[3].title = SeriesLabel(v="下限80%CI")

        # x軸
        ref_x = Reference(ws, min_col=COL_PERIOD, min_row=data_start, max_row=data_end)
        chart.set_categories(ref_x)

        charts.append((chart, base_row))

    # チャートを埋め込む（各ブロック右側 L列から）
    chart_col = get_column_letter(N_COLS + 2)
    for chart, base_row in charts:
        ws.add_chart(chart, f"{chart_col}{base_row}")

    print("  C_MarketForecast シート構築完了")
    print(f"  成分数: {len(INGREDIENTS)}, 実績: {N_ACTUAL}ヶ月, 予測: {N_FORECAST}ヶ月")
    print(f"  FORECAST.ETS 季節性パラメータ: {SEASONALITY}")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def main():
    print(f"📂 対象ファイル: {XLSX}")
    wb = load_workbook(XLSX)
    print("① C_MarketForecast シートを構築...")
    build_c_market_forecast(wb)
    wb.save(XLSX)
    print(f"\n✅ 保存完了: {XLSX}")


if __name__ == "__main__":
    main()
