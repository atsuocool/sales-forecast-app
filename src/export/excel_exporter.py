"""Excel エクスポート: 月次予測・年次サマリー・シナリオ比較の 3 シート"""
import io
from datetime import date

import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter

from src.export._helpers import INGREDIENTS, load_sku_fc_df, next_period, get_actual_sellout_12m
from src.forecast.currency import convert_amounts


_HDR_FILL  = PatternFill("solid", fgColor="2563EB")
_HDR_FONT  = Font(bold=True, color="FFFFFF", name="Arial")
_TITLE_FONT = Font(bold=True, size=13, name="Arial")
_SUBHDR_FILL = PatternFill("solid", fgColor="DBEAFE")
_SUBHDR_FONT = Font(bold=True, name="Arial")
_CENTER = Alignment(horizontal="center")
_RIGHT  = Alignment(horizontal="right")


def _write_header(ws, row: int, cols: list):
    for c, (label, width) in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font = _HDR_FONT
        cell.fill = _HDR_FILL
        cell.alignment = _CENTER
        ws.column_dimensions[get_column_letter(c)].width = width


def _monthly_sheet(wb, df_adj, df_exc, currency: str, fx: float, ing_name: str):
    ws = wb.create_sheet("月次予測")
    ws.cell(1, 1, f"月次 SKU 予測 — {ing_name}").font = _TITLE_FONT
    ws.cell(2, 1, f"生成日: {date.today()}  通貨: {currency}  FX: {fx:.1f} JPY/USD").font = Font(size=9, italic=True)

    unit = "万JPY" if currency == "JPY" else "万USD"
    cols = [
        ("月", 10), ("SKU", 12), (f"加味 数量(本)", 14),
        (f"加味 金額({unit})", 16), (f"非加味 数量(本)", 14), (f"非加味 金額({unit})", 16),
        ("差分 金額", 14),
    ]
    _write_header(ws, 4, cols)

    df_adj = df_adj.copy()
    df_exc = df_exc.copy()
    for df in [df_adj, df_exc]:
        df["amount"] = convert_amounts(df["amount_jpy"].values, currency, fx) / 10_000

    merged = df_adj.merge(df_exc, on=["period", "sku_id"], suffixes=("_adj", "_exc"))

    r = 5
    for _, row_data in merged.iterrows():
        diff = row_data["amount_adj"] - row_data["amount_exc"]
        vals = [
            row_data["period"], row_data["sku_id"],
            int(row_data["units_adj"]), round(row_data["amount_adj"], 2),
            int(row_data["units_exc"]), round(row_data["amount_exc"], 2),
            round(diff, 2),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            if c in (3, 5):
                cell.number_format = "#,##0"
                cell.alignment = _RIGHT
            elif c in (4, 6, 7):
                cell.number_format = "#,##0.00"
                cell.alignment = _RIGHT
                if c == 7:
                    cell.font = Font(color="DC2626" if diff > 0 else "16A34A")
        r += 1

    ws.freeze_panes = "A5"


def _annual_sheet(wb, df_adj, currency: str, fx: float, ing_name: str):
    ws = wb.create_sheet("年次サマリー")
    ws.cell(1, 1, f"年次サマリー — {ing_name}").font = _TITLE_FONT
    unit = "万JPY" if currency == "JPY" else "万USD"

    df = df_adj.copy()
    df["year"] = df["period"].str[:4]
    df["amount"] = convert_amounts(df["amount_jpy"].values, currency, fx) / 10_000
    piv = df.groupby(["sku_id", "year"])[["units", "amount"]].sum().reset_index()

    years = sorted(piv["year"].unique())
    skus  = sorted(piv["sku_id"].unique())

    # Header
    ws.cell(3, 1, "SKU").font = _SUBHDR_FONT
    for ci, y in enumerate(years, 2):
        ws.cell(3, ci * 2,     f"{y} 数量(本)").font  = _SUBHDR_FONT
        ws.cell(3, ci * 2 + 1, f"{y} 金額({unit})").font = _SUBHDR_FONT
        ws.cell(3, ci * 2).fill = _SUBHDR_FILL
        ws.cell(3, ci * 2 + 1).fill = _SUBHDR_FILL

    ws.column_dimensions["A"].width = 14
    for i in range(2, 2 + len(years) * 2):
        ws.column_dimensions[get_column_letter(i)].width = 15

    r = 4
    for sku in skus:
        ws.cell(r, 1, sku)
        for ci, y in enumerate(years, 2):
            sub = piv[(piv["sku_id"] == sku) & (piv["year"] == y)]
            units = int(sub["units"].sum()) if not sub.empty else 0
            amt   = round(sub["amount"].sum(), 1) if not sub.empty else 0.0
            ws.cell(r, ci * 2,     units).number_format = "#,##0"
            ws.cell(r, ci * 2 + 1, amt).number_format   = "#,##0.0"
        r += 1

    # 合計行
    ws.cell(r, 1, "合計").font = Font(bold=True)
    for ci, y in enumerate(years, 2):
        sub = piv[piv["year"] == y]
        ws.cell(r, ci * 2,     int(sub["units"].sum())).number_format   = "#,##0"
        ws.cell(r, ci * 2 + 1, round(sub["amount"].sum(), 1)).number_format = "#,##0.0"
        ws.cell(r, ci * 2).font     = Font(bold=True)
        ws.cell(r, ci * 2 + 1).font = Font(bold=True)

    ws.freeze_panes = "B4"


def _scenario_sheet(wb, df_adj, df_exc, currency: str, fx: float, ing_name: str):
    ws = wb.create_sheet("シナリオ比較")
    ws.cell(1, 1, f"シナリオ比較（加味 vs 非加味）— {ing_name}").font = _TITLE_FONT
    unit = "万JPY" if currency == "JPY" else "万USD"

    for df in [df_adj, df_exc]:
        df["year"] = df["period"].str[:4]
        df["amount"] = convert_amounts(df["amount_jpy"].values, currency, fx) / 10_000

    adj_year = df_adj.groupby("year")["amount"].sum()
    exc_year = df_exc.groupby("year")["amount"].sum()
    years    = sorted(set(adj_year.index) & set(exc_year.index))

    cols = [("年度", 10), (f"加味({unit})", 16), (f"非加味({unit})", 16),
            (f"差分({unit})", 16), ("影響率(%)", 12)]
    _write_header(ws, 3, cols)

    r = 4
    for y in years:
        a, e = adj_year.get(y, 0), exc_year.get(y, 0)
        diff = a - e
        rate = f"{diff / e * 100:+.1f}%" if e else "—"
        ws.cell(r, 1, y)
        ws.cell(r, 2, round(a, 1)).number_format    = "#,##0.0"
        ws.cell(r, 3, round(e, 1)).number_format    = "#,##0.0"
        ws.cell(r, 4, round(diff, 1)).number_format = "#,##0.0"
        ws.cell(r, 4).font = Font(color="DC2626" if diff > 0 else "16A34A")
        ws.cell(r, 5, rate).alignment = _CENTER
        r += 1

    cumul = sum(adj_year.get(y, 0) - exc_year.get(y, 0) for y in years)
    ws.cell(r + 1, 1, f"36ヶ月累計影響額: {cumul:+,.1f} {unit}").font = Font(bold=True, size=12)


def generate_excel(
    conn, ing_id: str, axis1: str, axis2: str, currency: str, fx: float
) -> bytes:
    df_adj = load_sku_fc_df(conn, ing_id, "regulatory_adjusted", axis2)
    df_exc = load_sku_fc_df(conn, ing_id, "regulatory_excluded",  axis2)
    ing_name = INGREDIENTS.get(ing_id, ing_id)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    _monthly_sheet(wb, df_adj.copy(), df_exc.copy(), currency, fx, ing_name)
    _annual_sheet(wb, df_adj.copy(), currency, fx, ing_name)
    _scenario_sheet(wb, df_adj.copy(), df_exc.copy(), currency, fx, ing_name)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
